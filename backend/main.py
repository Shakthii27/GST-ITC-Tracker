from extractor import (
    extract_text_from_pdf_bytes, parse_invoices_from_text, parse_payments_from_text,
    _is_scanned, _call_claude_vision, _call_claude_text, _pdf_to_base64_images, INVOICE_SYSTEM
)
from xml.parsers.expat import errors

from fastapi import FastAPI, UploadFile, File, HTTPException, Request
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse
from typing import List
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

from extractor import extract_text_from_pdf_bytes, parse_invoices_from_text, parse_payments_from_text
from database import insert_invoices, insert_payments, get_all_invoices, get_all_payments, reset_all
from reconciler import reconcile
import asyncio
import time
import random
from anthropic import APIStatusError

app = FastAPI(title="Invoice Tracker API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.get("/")
def root():
    return {"status": "Invoice Tracker API running"}


@app.post("/upload/invoices")
async def upload_invoices(files: List[UploadFile] = File(...)):
    all_extracted = []
    errors = []
    for file in files:
        if not file.filename.lower().endswith(".pdf"):
            errors.append(f"{file.filename}: not a PDF")
            continue
        try:
            pdf_bytes = await file.read()
            pages = extract_text_from_pdf_bytes(pdf_bytes)
            # ↓ pass pdf_bytes here
            invoices = parse_invoices_from_text(pages, file.filename, pdf_bytes)
            if invoices:
                insert_invoices(invoices)
                all_extracted.extend(invoices)
            else:
                errors.append(f"{file.filename}: no invoices detected")
        except Exception as e:
            errors.append(f"{file.filename}: {str(e)}")
    for invoice in all_extracted:
        if "_id" in invoice:
            invoice["_id"] = str(invoice["_id"])
    return {"inserted": len(all_extracted), "invoices": all_extracted, "errors": errors}


@app.post("/upload/audit")
async def upload_audit(files: List[UploadFile] = File(...)):
    all_extracted = []
    errors = []
    for file in files:
        if not file.filename.lower().endswith(".pdf"):
            errors.append(f"{file.filename}: not a PDF")
            continue
        try:
            pdf_bytes = await file.read()
            pages = extract_text_from_pdf_bytes(pdf_bytes)
            # ↓ pass pdf_bytes here
            payments = parse_payments_from_text(pages, file.filename, pdf_bytes)
            if payments:
                insert_payments(payments)
                all_extracted.extend(payments)
            else:
                errors.append(f"{file.filename}: no payments detected")
        except Exception as e:
            errors.append(f"{file.filename}: {str(e)}")
    for payment in all_extracted:
        if "_id" in payment:
            payment["_id"] = str(payment["_id"])
    return {"inserted": len(all_extracted), "payments": all_extracted, "errors": errors}


@app.get("/reconcile")
def get_reconciliation():
    """Return the 3-list reconciliation result."""
    return reconcile()


@app.get("/invoices")
def list_invoices():
    return {"invoices": get_all_invoices()}


@app.get("/payments")
def list_payments():
    return {"payments": get_all_payments()}


@app.delete("/reset")
def reset():
    reset_all()
    return {"message": "All data cleared"}

MAX_REQUESTS_PER_WINDOW = 5
WINDOW_SECONDS = 60

_lock = asyncio.Lock()
_request_times: list[float] = []


async def _wait_for_slot():
    async with _lock:
        now = time.monotonic()
        global _request_times
        _request_times = [t for t in _request_times if now - t < WINDOW_SECONDS]

        if len(_request_times) >= MAX_REQUESTS_PER_WINDOW:
            sleep_for = WINDOW_SECONDS - (now - _request_times[0]) + 0.25
        else:
            sleep_for = 0

        if sleep_for > 0:
            await asyncio.sleep(sleep_for)
            now = time.monotonic()
            _request_times = [t for t in _request_times if now - t < WINDOW_SECONDS]

        _request_times.append(time.monotonic())


async def rate_limited_claude_call(claude_client, **kwargs):
    max_retries = 5
    base_delay = 2.0

    for attempt in range(max_retries):
        await _wait_for_slot()
        try:
            return await asyncio.to_thread(claude_client.messages.create, **kwargs)
        except APIStatusError as e:
            if e.status_code == 429 and attempt < max_retries - 1:
                retry_after = None
                try:
                    retry_after = float(e.response.headers.get("retry-after"))
                except Exception:
                    pass
                delay = retry_after if retry_after else base_delay * (2 ** attempt) + random.uniform(0, 1)
                await asyncio.sleep(delay)
                continue
            raise

from anthropic import Anthropic

claude_client = Anthropic()

@app.post("/api/claude")
async def claude_proxy(request: Request):
    body = await request.json()
    try:
        response = await rate_limited_claude_call(
            claude_client,
            model=body["model"],
            max_tokens=body.get("max_tokens", 4096),
            system=body.get("system", ""),
            messages=body["messages"],
        )
        return {"content": [{"type": "text", "text": response.content[0].text}]}
    except APIStatusError as e:
        raise HTTPException(status_code=e.status_code, detail=str(e))

@app.get("/export/excel")
def export_excel():
    """Export reconciliation result as Excel file."""
    data = reconcile()
    wb = openpyxl.Workbook()

    header_font = Font(bold=True, color="FFFFFF")
    red_fill = PatternFill("solid", fgColor="C0392B")
    green_fill = PatternFill("solid", fgColor="27AE60")
    blue_fill = PatternFill("solid", fgColor="2980B9")

    def make_sheet(ws, records, fill, columns):
        ws.append(columns)
        for cell in ws[1]:
            cell.font = header_font
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")
        for r in records:
            if columns[0] == "Vendor":
                if r["status"] == "unpaid":
                    ws.append([r["vendor_name"], r["gstin"],
                               r["total_invoiced"], r["total_paid"], r["pending"]])
                elif r["status"] == "overpaid":
                    ws.append([r["vendor_name"], r["gstin"],
                               r["total_invoiced"], r["total_paid"], r["excess"]])
                else:
                    ws.append([r["vendor_name"], r["gstin"],
                               r["total_invoiced"], r["total_paid"], "—"])
        for col in ws.columns:
            max_len = max(len(str(c.value or "")) for c in col)
            ws.column_dimensions[col[0].column_letter].width = max_len + 4

    ws1 = wb.active
    ws1.title = "Not Paid"
    make_sheet(ws1, data["unpaid"], red_fill,
               ["Vendor", "GSTIN", "Total Invoiced (₹)", "Paid (₹)", "Pending (₹)"])

    ws2 = wb.create_sheet("Fully Paid")
    make_sheet(ws2, data["paid"], green_fill,
               ["Vendor", "GSTIN", "Total Invoiced (₹)", "Paid (₹)", "Pending (₹)"])

    ws3 = wb.create_sheet("Overpaid")
    make_sheet(ws3, data["overpaid"], blue_fill,
               ["Vendor", "GSTIN", "Total Invoiced (₹)", "Paid (₹)", "Excess (₹)"])

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)

    filename = f"reconciliation_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return StreamingResponse(
        stream,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )

@app.post("/debug/pdf")
async def debug_pdf(file: UploadFile = File(...)):
    pdf_bytes = await file.read()
    pages = extract_text_from_pdf_bytes(pdf_bytes)
    is_scanned = _is_scanned(pages)
    
    result = {
        "is_scanned": is_scanned,
        "page_count": len(pages),
        "text_preview": [p[:200] for p in pages],
        "total_chars": sum(len(p) for p in pages)
    }
    
    if is_scanned:
        # Try vision and show raw Claude response
        from pdf2image import convert_from_bytes
        import base64, io
        b64_images = _pdf_to_base64_images(pdf_bytes)
        result["image_count"] = len(b64_images)
        raw = _call_claude_vision(INVOICE_SYSTEM, b64_images)
        result["claude_raw_response"] = raw
    else:
        full_text = "\n\n--- PAGE BREAK ---\n\n".join(pages).strip()
        raw = _call_claude_text(INVOICE_SYSTEM, f"Extract all invoices:\n\n{full_text}")
        result["claude_raw_response"] = raw
        result["full_text_sent"] = full_text[:500]
    
    return result